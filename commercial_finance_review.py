"""
commercial_finance_review.py

A validation agent that role-plays a senior commercial real-estate underwriter.
It (1) generates several realistic commercial deals, (2) enters those exact
numbers into the REAL Commercial Real Estate Deal Analyzer workbook, (3) forces
a full LibreOffice recalculation so the tool's own live formulas produce NOI /
cap rate / DSCR / price-per-SqFt / verdict, then (4) puts the underwriter persona
back on to critique whether each computed metric and Buy/Negotiate/Pass call is
what a real CRE finance pro would conclude -- flagging any assumption a
professional would dispute (the NNN 5% op-ex factor, the 1.25x DSCR floor, the
benchmark logic, etc.).

This closes the one honest gap: no human commercial-finance review of the
assumptions. The persona IS that review, and it is grounded in the tool's
actual recalculated output, not a guess.

Run: python commercial_finance_review.py
Requires LibreOffice (soffice.exe) and ANTHROPIC_API_KEY (.env).
"""
import json
import os
import re
import shutil
import tempfile

from openpyxl import load_workbook

import creator
import simulate_buyer as sb
from build_commercial import COMMERCIAL

FNAME = creator.safe_filename(COMMERCIAL["suggested_etsy_title"])
XLSX = os.path.join(creator.PRODUCTS_DIR, f"{FNAME}.xlsx")
SHEET = "Deal Comparison"
FIRST_ROW = 6           # deal rows start here (row 5 is the header)
REPORT = os.path.join(creator.PRODUCTS_DIR, "commercial_finance_review_report.txt")

# Column letter for each input field, in the workbook's own order.
INPUT_COLS = [
    ("A", "property_name"),
    ("B", "asset_type"),
    ("C", "square_footage"),
    ("D", "purchase_price"),
    ("E", "gross_rental_income"),
    ("F", "vacancy_pct"),
    ("G", "operating_expenses"),
    ("H", "capital_reserve_per_sqft"),
    ("I", "management_fee_pct"),
    ("J", "lease_type"),
    ("K", "loan_amount"),
    ("L", "interest_rate"),
    ("M", "amortization_years"),
    ("N", "target_dscr"),
    ("O", "benchmark_price_per_sqft"),
    ("P", "investor_goal"),
]

# Computed columns we read back after recalculation.
OUTPUT_COLS = [
    ("S", "Capital Reserve"),
    ("T", "Management Fee"),
    ("U", "NOI"),
    ("V", "Cap Rate %"),
    ("W", "Price Per SqFt"),
    ("X", "Price vs Benchmark %"),
    ("Y", "Loan Constant %"),
    ("Z", "Annual Debt Service"),
    ("AA", "DSCR"),
    ("AB", "Cash Invested"),
    ("AC", "Cash-on-Cash Return %"),
    ("AD", "Break-Even Occupancy %"),
    ("AE", "Verdict"),
]

UNDERWRITER_PERSONA = (
    "You are a senior commercial real estate credit underwriter with 20 years at "
    "a regional bank and a life-insurance lender. You have personally sized and "
    "approved several hundred million dollars of retail, office, industrial, and "
    "mixed-use loans. You know cold: NOI = effective gross income minus operating "
    "expenses; cap rate = NOI / price; DSCR = NOI / annual debt service and that "
    "most lenders floor DSCR at 1.20x-1.30x; loan constant; cash-on-cash; "
    "price-per-SqFt benchmarking against a submarket median; and how NNN vs "
    "Modified Gross vs Full-Service Gross leases shift operating-expense burden "
    "between landlord and tenant. You are skeptical, precise, and you call out "
    "anything that would not survive a real credit committee."
)

DEAL_GEN_INSTRUCTIONS = """\
Generate {n} realistic, varied commercial real-estate deals that a real investor
would actually bring to an underwriter. Make them genuinely diverse so they
exercise different verdicts:
  - one clean, well-financed deal that should clearly underwrite (a "Buy")
  - one where DSCR is tight / below ~1.25x (financing won't pencil)
  - one priced above its submarket benchmark price-per-SqFt
  - one value-add / higher-vacancy or thin-cash-flow deal

Use real-world numbers consistent with each asset class. Vary lease type across
NNN, Modified Gross, and Full-Service Gross. Vary investor goal between
"Cash Flow" and "Appreciation".

Return STRICT JSON: an object with key "deals" -> array of exactly {n} objects.
Each object MUST have these keys with these types:
  property_name (str), asset_type (one of Retail/Office/Industrial/Mixed-Use),
  square_footage (int), purchase_price (int), gross_rental_income (int),
  vacancy_pct (float, e.g. 0.06 for 6%), operating_expenses (int),
  capital_reserve_per_sqft (float, replacement/CapEx reserve in $/SqFt, e.g.
  0.20 retail, 0.25 office, 0.15 industrial),
  management_fee_pct (float of effective gross income, e.g. 0.04),
  lease_type (one of "NNN","Modified Gross","Full-Service Gross"),
  loan_amount (int), interest_rate (float, e.g. 0.072), amortization_years (int),
  target_dscr (float, e.g. 1.25), benchmark_price_per_sqft (int),
  investor_goal (one of "Cash Flow","Appreciation"),
  expected_call (str: your own pre-tool gut call -- "Buy","Negotiate", or "Pass",
  with a one-sentence reason). expected_call is your prediction BEFORE seeing the
  tool's output; it lets us check whether the tool agrees with a human pro.
Return ONLY the JSON object."""


def generate_deals(client, n=4):
    user = DEAL_GEN_INSTRUCTIONS.format(n=n)
    data = creator.call_claude_json(client, UNDERWRITER_PERSONA, user, max_tokens=4096)
    deals = data.get("deals", [])
    if not deals:
        raise SystemExit("[FATAL] deal generation returned no deals")
    return deals


def write_deals_into_workbook(deals, dest):
    """Copy the real analyzer and overwrite its sample deal rows with our deals,
    clearing any leftover sample rows so only our deals compute."""
    wb = load_workbook(XLSX)
    ws = wb[SHEET]
    # find how many sample rows existed (rows with a Property Name in col A)
    existing = 0
    r = FIRST_ROW
    while ws[f"A{r}"].value not in (None, ""):
        existing += 1
        r += 1
    # write our deals
    for i, deal in enumerate(deals):
        row = FIRST_ROW + i
        for col, key in INPUT_COLS:
            ws[f"{col}{row}"] = deal[key]
    # clear any sample input rows beyond our deals (leave formulas alone)
    for i in range(len(deals), existing):
        row = FIRST_ROW + i
        for col, _ in INPUT_COLS:
            ws[f"{col}{row}"] = None
    wb.save(dest)
    wb.close()


def read_outputs(recalced, n):
    wb = load_workbook(recalced, data_only=True)
    ws = wb[SHEET]
    results = []
    for i in range(n):
        row = FIRST_ROW + i
        out = {}
        for col, label in OUTPUT_COLS:
            out[label] = ws[f"{col}{row}"].value
        results.append(out)
    wb.close()
    return results


def fmt_metric(label, v):
    if v is None or v == "":
        return f"{label}: (blank)"
    if isinstance(v, float):
        if "%" in label or "Return" in label or "Rate" in label or "Benchmark" in label or "Occupancy" in label:
            return f"{label}: {v*100:.2f}%"
        if "DSCR" in label or "Constant" in label:
            return f"{label}: {v:.3f}"
        return f"{label}: {v:,.0f}"
    return f"{label}: {v}"


def call_json_robust(client, system, user, max_tokens):
    """Like creator.call_claude_json but tolerant of prose/fences around the JSON
    object -- extracts the first balanced {...} block."""
    resp = client.messages.create(
        model=creator.MODEL_NAME, max_tokens=max_tokens,
        system=system, messages=[{"role": "user", "content": user}],
    )
    text = "".join(b.text for b in resp.content if hasattr(b, "text")).strip()
    if text.startswith("```"):
        text = text.strip("`")
        if text.lower().startswith("json"):
            text = text[4:]
    text = text.strip()
    try:
        return creator.sanitize_text(json.loads(text))
    except json.JSONDecodeError:
        m = re.search(r"\{.*\}", text, re.DOTALL)
        if not m:
            raise
        return creator.sanitize_text(json.loads(m.group(0)))


def critique(client, deal, result):
    tool_block = "\n".join(fmt_metric(l, result.get(l)) for _, l in OUTPUT_COLS)
    user = f"""\
You entered this real deal into a commercial RE analysis tool a client of yours
is considering buying. Here are the EXACT inputs and the tool's OWN computed
output (its live formulas, freshly recalculated).

DEAL INPUTS:
{json.dumps({k: deal[k] for _, k in INPUT_COLS}, indent=2)}

Your pre-tool gut call was: {deal.get('expected_call', '(none)')}

THE TOOL COMPUTED:
{tool_block}

For reference, the tool now models NOI = effective gross income - landlord
operating expenses - a replacement/CapEx reserve (Capital Reserve $/SqFt x
square footage) - a management fee (% of effective gross income). Its lease
logic charges the landlord 5% of operating expenses on NNN, 50% on Modified
Gross, and 100% on Full-Service Gross. Break-even occupancy includes the
reserve, management fee, and debt service. Its verdict flags Pass/Negotiate if
DSCR < target or break-even occupancy > 95%, considers price-vs-benchmark > 5%,
flags Negotiate on any deal with vacancy > 15% (lease-up risk) regardless of
goal, and for Cash-Flow-goal deals flags Negotiate if cash-on-cash < 6%;
otherwise Buy.

As the underwriter, assess STRICTLY in JSON with keys:
  math_correct (bool): are NOI, cap rate, DSCR, price/SqFt, cash-on-cash
    arithmetically right given the inputs?
  verdict_sound (bool): is the Buy/Negotiate/Pass call what you'd tell the client?
  agrees_with_your_gut (bool): does the tool's verdict match your pre-tool call?
  disputed_assumptions (array of strings): any assumption a real underwriter
    would push back on (e.g. the NNN 5% op-ex factor, the DSCR floor, benchmark
    handling). Empty array if none.
  verdict_quote (str): the tool's verdict text.
  underwriter_note (str): 1-3 sentences, your professional read of this deal and
    whether the tool served the client well.
Return ONLY the JSON object."""
    return call_json_robust(client, UNDERWRITER_PERSONA, user, max_tokens=1536)


def main():
    if not os.path.exists(sb.SOFFICE):
        raise SystemExit(f"[FATAL] LibreOffice not found at {sb.SOFFICE}")
    if not os.path.exists(XLSX):
        raise SystemExit(f"[FATAL] analyzer workbook missing: {XLSX}")

    client = creator.load_anthropic_client()
    n = 4
    print(f"[*] Underwriter persona generating {n} realistic deals (Opus)...\n")
    deals = generate_deals(client, n)
    for d in deals:
        print(f"    - {d['property_name']} ({d['asset_type']}, {d['lease_type']}, "
              f"goal {d['investor_goal']}) | gut: {d.get('expected_call','')}")

    tmp = tempfile.mkdtemp(prefix="comm_review_")
    profile = os.path.join(tmp, "prof")
    staged = os.path.join(tmp, "review_deals.xlsx")
    write_deals_into_workbook(deals, staged)

    print("\n[*] Recalculating the analyzer's live formulas in LibreOffice...")
    proc = sb.start_soffice(profile)
    try:
        print("   ", sb.batch_recalc(tmp))
    finally:
        try: proc.terminate()
        except Exception: pass

    results = read_outputs(staged, n)
    shutil.rmtree(tmp, ignore_errors=True)

    print("\n[*] Underwriter reviewing each computed result...\n")
    lines = []
    def out(s=""):
        print(s); lines.append(s)

    out("=" * 72)
    out("COMMERCIAL FINANCE PRO -- VALIDATION REVIEW")
    out("Senior CRE underwriter persona vs. the live analyzer output")
    out("=" * 72)

    math_ok = verdict_ok = agree = 0
    all_disputes = []
    for i, (deal, result) in enumerate(zip(deals, results), 1):
        c = critique(client, deal, result)
        math_ok += bool(c.get("math_correct"))
        verdict_ok += bool(c.get("verdict_sound"))
        agree += bool(c.get("agrees_with_your_gut"))
        all_disputes += c.get("disputed_assumptions", [])

        out(f"\n[{i}] {deal['property_name']}  ({deal['asset_type']}, "
            f"{deal['lease_type']}, goal {deal['investor_goal']})")
        out(f"    Price ${deal['purchase_price']:,}  Loan ${deal['loan_amount']:,}  "
            f"Rate {deal['interest_rate']*100:.2f}%")
        for _, label in OUTPUT_COLS:
            out("      " + fmt_metric(label, result.get(label)))
        out(f"    underwriter gut call : {deal.get('expected_call','')}")
        out(f"    math correct         : {c.get('math_correct')}")
        out(f"    verdict sound        : {c.get('verdict_sound')}")
        out(f"    agrees with gut      : {c.get('agrees_with_your_gut')}")
        disputes = c.get("disputed_assumptions", [])
        out(f"    disputed assumptions : {disputes if disputes else 'none'}")
        out(f"    underwriter note     : {c.get('underwriter_note','')}")

    out("\n" + "=" * 72)
    out(f"math correct      : {math_ok}/{n} deals")
    out(f"verdict sound     : {verdict_ok}/{n} deals")
    out(f"agrees with pro   : {agree}/{n} deals")
    uniq = sorted(set(all_disputes))
    out(f"distinct assumptions flagged by the underwriter: {len(uniq)}")
    for u in uniq:
        out(f"   - {u}")
    verdict = ("PROFESSIONALLY SOUND" if math_ok == n and verdict_ok == n
               else "REVIEW THE FLAGGED ITEMS")
    out("=" * 72)
    out(f"RESULT: {verdict}")
    out("=" * 72)

    with open(REPORT, "w", encoding="utf-8") as f:
        f.write("\n".join(lines) + "\n")
    print(f"\n[saved] {REPORT}")


if __name__ == "__main__":
    main()
