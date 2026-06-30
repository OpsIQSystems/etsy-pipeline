"""
simulate_use.py  --  "Simulated use of the product" hero scene generator.

Reads a REAL product workbook, recalculates it (LibreOffice UNO bridge) to get
genuine computed values, then renders an animated phone-style mini-app of the
tool in use: the actual input fields fill in with real sample numbers, the math
runs, and the real verdict/recommendation reveals in green. This is the
highest-converting UGC format -- it shows the product actually working.

Exposes:
  demo_data(xlsx_path) -> dict(display, fields, metric, verdict)
  build_use_frames(page, data, tmpdir, prefix) -> list[(png, dur, zoom)]
Standalone test:  python simulate_use.py <opportunity_index>
"""
import os
import shutil
import sys
import tempfile

from openpyxl import load_workbook

import creator
import simulate_buyer as sb
import ugc_engine as ug

SKIP_HEADER_WORDS = ("name", "type", "id", "date", "notes")
METRIC_HINTS = ("cost", "profit", "save", "total", "price", "revenue", "margin", "roi", "cash")


def _looks_like_verdict(v):
    return isinstance(v, str) and len(v.strip()) >= 8 and any(c.isalpha() for c in v) \
        and v.strip() not in sb.ERROR_VALUES


def demo_data(xlsx_path):
    """Recalc the workbook (own soffice session) and extract demo data."""
    tmp = tempfile.mkdtemp(prefix="use_")
    try:
        work = os.path.join(tmp, "p.xlsx")
        shutil.copyfile(xlsx_path, work)
        proc = sb.start_soffice(os.path.join(tmp, "prof"))
        try:
            sb.batch_recalc(tmp)
        finally:
            try:
                proc.terminate()
            except Exception:
                pass
        return extract_from_pair(xlsx_path, work)
    finally:
        shutil.rmtree(tmp, ignore_errors=True)


def extract_from_pair(formula_path, value_path):
    """Pure extraction: given the original (formulas) and a recalculated copy
    (values), pull real input fields + verdict + key metric. No soffice here --
    caller is responsible for having recalculated `value_path`."""
    if True:
        formula_wb = load_workbook(formula_path)             # formula-vs-input structure
        value_wb = load_workbook(value_path, data_only=True)  # recalculated values

        # decision sheet = most formulas
        best, bestn = None, -1
        for ws in formula_wb.worksheets:
            n = sum(1 for r in ws.iter_rows() for c in r
                    if isinstance(c.value, str) and c.value.startswith("="))
            if n > bestn:
                best, bestn = ws.title, n
        fws, vws = formula_wb[best], value_wb[best]

        # header row = first row with >=3 plain-string cells
        header_row = None
        for r in range(1, 14):
            strs = [c.value for c in fws[r]
                    if isinstance(c.value, str) and not c.value.startswith("=")]
            if len(strs) >= 3:
                header_row = r
                break
        if header_row is None:
            header_row = 1
        data_row = header_row + 1

        headers = {}  # col_letter -> header text
        for c in fws[header_row]:
            if isinstance(c.value, str) and c.value.strip():
                headers[c.column_letter] = c.value.strip()

        fields, metric, verdict = [], None, None
        for col, htext in headers.items():
            fcell = fws[f"{col}{data_row}"].value
            vcell = vws[f"{col}{data_row}"].value
            is_formula = isinstance(fcell, str) and fcell.startswith("=")
            low = htext.lower()
            if not is_formula:
                # candidate input field (skip names/types/ids)
                if isinstance(fcell, (int, float)) and not any(w in low for w in SKIP_HEADER_WORDS):
                    fields.append((htext, fcell))
            else:
                if verdict is None and _looks_like_verdict(vcell):
                    verdict = (htext, vcell.strip())
                elif metric is None and isinstance(vcell, (int, float)) \
                        and any(h in low for h in METRIC_HINTS):
                    metric = (htext, vcell)
        # fallbacks
        if verdict is None:
            # any computed string anywhere on the row
            for col in headers:
                vcell = vws[f"{col}{data_row}"].value
                if _looks_like_verdict(vcell):
                    verdict = (headers[col], vcell.strip()); break
        fields = fields[:4]
        return {"display": best, "fields": fields, "metric": metric, "verdict": verdict}


def _fmt(v):
    if isinstance(v, (int, float)):
        if abs(v) >= 100 or float(v).is_integer():
            return f"{v:,.0f}"
        return f"{v:,.2f}"
    return str(v)


def _app_html(display, fields, filled, caption, metric=None, verdict=None,
              metric_on=False, verdict_on=False):
    """Render the mini-app at a given animation state. `filled` = number of input
    fields shown filled so far."""
    rows = ""
    for i, (label, value) in enumerate(fields):
        shown = i < filled
        active = (i == filled - 1)
        box = (f'<span class="val">{_fmt(value)}</span>' if shown
               else '<span class="cursor">|</span>' if active else '')
        rows += (f'<div class="row{" active" if active else ""}">'
                 f'<span class="lbl">{label}</span>'
                 f'<span class="field">{box}</span></div>')
    result = ""
    if metric and metric_on:
        result += (f'<div class="metric"><span>{metric[0]}</span>'
                   f'<b>${_fmt(metric[1])}</b></div>')
    if verdict and verdict_on:
        vtxt = verdict[1]
        result += (f'<div class="verdict"><div class="vk">RECOMMENDATION</div>'
                   f'<div class="vt">{vtxt}</div></div>')
    inner = f"""
.app{{width:880px;background:#fff;border-radius:40px;overflow:hidden;
  box-shadow:0 40px 120px rgba(0,0,0,.5);border:1px solid rgba(255,255,255,.1);}}
.bar{{background:{ug.PALETTE['navy']};color:#fff;padding:34px 42px;display:flex;
  align-items:center;gap:18px;font-size:40px;font-weight:800;}}
.dot{{width:22px;height:22px;border-radius:50%;background:{ug.PALETTE['accent']};
  box-shadow:0 0 0 8px rgba(61,214,176,.25);}}
.body{{padding:30px 42px 46px;color:{ug.PALETTE['navy']};}}
.row{{display:flex;justify-content:space-between;align-items:center;
  padding:26px 4px;border-bottom:2px solid #EAEEF4;font-size:40px;}}
.row.active{{background:#F3FBF8;border-radius:14px;}}
.lbl{{color:#54657E;font-weight:600;}}
.field{{min-width:230px;text-align:right;}}
.val{{font-weight:900;color:{ug.PALETTE['navy']};}}
.cursor{{color:{ug.PALETTE['accent2']};font-weight:900;}}
.metric{{display:flex;justify-content:space-between;font-size:42px;
  margin-top:34px;color:#54657E;font-weight:700;}}
.metric b{{color:{ug.PALETTE['navy']};font-size:48px;}}
.verdict{{margin-top:30px;background:{ug.PALETTE['accent']};border-radius:24px;
  padding:34px 36px;color:{ug.PALETTE['navy']};}}
.vk{{font-size:30px;letter-spacing:5px;font-weight:900;opacity:.7;}}
.vt{{font-size:46px;font-weight:900;line-height:1.2;margin-top:10px;}}
.cap{{font-size:60px;font-weight:900;color:{ug.PALETTE['text']};text-align:center;
  margin-bottom:50px;line-height:1.15;padding:0 60px;}}
"""
    body = (f'<div class="cap">{caption}</div>'
            f'<div class="app"><div class="bar"><span class="dot"></span>{display}</div>'
            f'<div class="body">{rows}{result}</div></div>')
    return ug._frame(inner, body)


def build_use_frames(page, data, tmpdir, prefix):
    fields = data["fields"]
    metric = data.get("metric")
    verdict = data.get("verdict")
    states = []
    # type the inputs in
    for i in range(len(fields) + 1):
        states.append((_app_html(data["display"], fields, i,
                                 "Enter your real numbers", metric, verdict),
                       0.7, False))
    # math runs (metric appears)
    if metric:
        states.append((_app_html(data["display"], fields, len(fields),
                                 "It runs the math instantly", metric, verdict,
                                 metric_on=True), 1.0, False))
    # verdict reveal (held longer)
    if verdict:
        states.append((_app_html(data["display"], fields, len(fields),
                                 "And tells you what to do", metric, verdict,
                                 metric_on=bool(metric), verdict_on=True), 2.4, False))
    out = []
    for i, (html, dur, zoom) in enumerate(states):
        png = os.path.join(tmpdir, f"{prefix}_use{i}.png")
        ug.render_scene(page, html, png)
        out.append((png, dur, zoom))
    return out


if __name__ == "__main__":
    import json
    from playwright.sync_api import sync_playwright
    idx = int(sys.argv[1]) if len(sys.argv) > 1 else 8
    with open(creator.INPUT_FILE, encoding="utf-8") as f:
        opp = json.load(f)[idx]
    xlsx = os.path.join(creator.PRODUCTS_DIR,
                        f"{creator.safe_filename(opp['suggested_etsy_title'])}.xlsx")
    print("extracting real demo data (recalculating)...")
    data = demo_data(xlsx)
    print("display sheet:", data["display"])
    print("fields:", data["fields"])
    print("metric:", data["metric"])
    print("verdict:", data["verdict"])
    out_dir = os.path.join(creator.PRODUCTS_DIR, "videos", "_usetest")
    os.makedirs(out_dir, exist_ok=True)
    with sync_playwright() as p:
        b = p.chromium.launch(headless=True, channel="msedge")
        pg = b.new_page(viewport={"width": ug.WIDTH, "height": ug.HEIGHT}, device_scale_factor=1)
        frames = build_use_frames(pg, data, out_dir, "demo")
        ug.assemble(frames, os.path.join(out_dir, "demo.mp4"))
        # also save the verdict frame for quick inspection
        shutil.copyfile(frames[-1][0], os.path.join(out_dir, "_verdict_frame.png"))
        b.close()
    print("built", os.path.join(out_dir, "demo.mp4"))
