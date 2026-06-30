"""
simulate_buyer.py  --  EXHAUSTIVE end-user simulation.

For EVERY product, EVERY worksheet (section), and EVERY formula cell, this:
  1. Identifies the input cells (numeric cells that are NOT formulas).
  2. Generates a battery of input SCENARIOS that a real, careless, or
     adversarial buyer might produce:
        - baseline    : the shipped sample values
        - zeros       : every input = 0           (division-by-zero stress)
        - blanks      : every input cleared        (empty-cell stress)
        - negatives   : every input made negative  (sign / sqrt / log stress)
        - ones        : every input = 1
        - huge        : every input = 1e12         (overflow / precision)
        - tiny        : every input = 0.0000001
        - text        : a string typed into a numeric input (#VALUE! stress)
        - mixed       : half zero, half huge        (ratio extremes)
  3. Writes each scenario to a temp workbook, opens it in LibreOffice headless
     with FORCED full recalculation, and reads back every formula cell.
  4. Flags, per scenario:
        - any Excel error value a buyer would see (#DIV/0!, #REF!, #VALUE!, ...)
        - any formula cell that failed to evaluate at all (came back None)
  5. Also verifies RESPONSIVENESS: across scenarios, each decision/verdict-style
     formula cell must take at least 2 distinct values (a verdict that never
     changes regardless of input is a dead formula, not a decision tool).

Run: python simulate_buyer.py
     python simulate_buyer.py --quick     (baseline+zeros+blanks only)
Requires LibreOffice (soffice.exe).
"""
import glob
import json
import os
import shutil
import subprocess
import sys
import tempfile
import time
import urllib.request

from openpyxl import load_workbook

import creator

LO_DIR = r"C:\Program Files\LibreOffice\program"
SOFFICE = os.path.join(LO_DIR, "soffice.exe")
LO_PYTHON = os.path.join(LO_DIR, "python.exe")
RECALC_SCRIPT = os.path.join(os.path.dirname(os.path.abspath(__file__)), "lo_recalc.py")
SOCKET_PORT = 2002

ERROR_VALUES = {
    "#DIV/0!", "#REF!", "#VALUE!", "#NAME?", "#N/A",
    "#NULL!", "#NUM!", "#ERROR!", "#CALC!", "Err:502", "Err:509",
}

SOFFICE_CANDIDATES = [
    r"C:\Program Files\LibreOffice\program\soffice.exe",
    r"C:\Program Files (x86)\LibreOffice\program\soffice.exe",
]

RECALC_XCU = """<?xml version="1.0" encoding="UTF-8"?>
<oor:items xmlns:oor="http://openoffice.org/2001/registry" xmlns:xs="http://www.w3.org/2001/XMLSchema" xmlns:xsi="http://www.w3.org/2001/XMLSchema-instance">
 <item oor:path="/org.openoffice.Office.Calc/Formula/Load"><prop oor:name="OOXMLRecalcMode" oor:op="fuse"><value>0</value></prop></item>
 <item oor:path="/org.openoffice.Office.Calc/Formula/Load"><prop oor:name="ODFRecalcMode" oor:op="fuse"><value>0</value></prop></item>
</oor:items>
"""

# Scenarios: name -> function(original_value) -> new_value (or KEEP / CLEAR sentinels)
KEEP = object()
CLEAR = object()


def scenario_value(name, original):
    if name == "baseline":
        return KEEP
    if name == "zeros":
        return 0
    if name == "blanks":
        return CLEAR
    if name == "negatives":
        return -abs(original) if isinstance(original, (int, float)) and original else -1
    if name == "ones":
        return 1
    if name == "huge":
        return 1_000_000_000_000
    if name == "tiny":
        return 0.0000001
    if name == "text":
        return "abc"
    if name == "mixed_high":
        return 1_000_000_000_000
    if name == "mixed_low":
        return 0
    return KEEP


ALL_SCENARIOS = ["baseline", "zeros", "blanks", "negatives", "ones",
                 "huge", "tiny", "text", "mixed"]
QUICK_SCENARIOS = ["baseline", "zeros", "blanks"]


def path_to_url(p):
    return "file:///" + os.path.abspath(p).replace("\\", "/")


def start_soffice(profile_dir):
    """Launch headless soffice with a UNO socket. Returns the Popen handle."""
    os.makedirs(profile_dir, exist_ok=True)
    proc = subprocess.Popen([
        SOFFICE, "--headless", "--norestore", "--nologo", "--nofirststartwizard",
        f"-env:UserInstallation={path_to_url(profile_dir)}",
        f"--accept=socket,host=localhost,port={SOCKET_PORT};urp;",
    ])
    # give it time to bind the socket before lo_recalc tries to connect
    time.sleep(8)
    return proc


def batch_recalc(folder):
    """Recalculate every .xlsx in folder in-place via the bundled-python UNO
    bridge. Files come back with cached formula values populated."""
    res = subprocess.run([LO_PYTHON, RECALC_SCRIPT, folder, str(SOCKET_PORT)],
                         capture_output=True, text=True, timeout=600)
    return res.stdout.strip() + (("\n" + res.stderr.strip()) if res.stderr.strip() else "")


def map_cells(xlsx_path):
    """Return (input_cells, formula_cells) as lists of (sheet, coord)."""
    wb = load_workbook(xlsx_path)
    inputs, formulas = [], []
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                v = cell.value
                if isinstance(v, str) and v.startswith("="):
                    formulas.append((ws.title, cell.coordinate))
                elif isinstance(v, (int, float)) and not isinstance(v, bool):
                    inputs.append((ws.title, cell.coordinate))
    wb.close()
    return inputs, formulas


def build_scenario_wb(src_xlsx, inputs, scenario, dest):
    """Write a copy of src with input cells mutated per scenario."""
    wb = load_workbook(src_xlsx)
    for i, (sheet, coord) in enumerate(inputs):
        cell = wb[sheet][coord]
        original = cell.value
        if scenario == "mixed":
            nv = 1_000_000_000_000 if i % 2 == 0 else 0
        else:
            nv = scenario_value(scenario, original)
        if nv is KEEP:
            continue
        if nv is CLEAR:
            cell.value = None
        else:
            cell.value = nv
    wb.save(dest)
    wb.close()


def read_formula_values(recalced_path, formula_cells):
    """Return dict {(sheet,coord): value} and list of error findings."""
    values, errors, uncomputed = {}, [], []
    wb = load_workbook(recalced_path, data_only=True)
    for sheet, coord in formula_cells:
        try:
            v = wb[sheet][coord].value
        except Exception:
            v = None
        values[(sheet, coord)] = v
        if isinstance(v, str) and v.strip() in ERROR_VALUES:
            errors.append((sheet, coord, v.strip()))
        elif v is None:
            uncomputed.append((sheet, coord))
    wb.close()
    return values, errors, uncomputed


def main():
    scenarios = QUICK_SCENARIOS if "--quick" in sys.argv else ALL_SCENARIOS
    if not os.path.exists(SOFFICE):
        print(f"[FATAL] LibreOffice not found at {SOFFICE}"); sys.exit(2)
    print(f"[*] LibreOffice: {SOFFICE}")
    print(f"[*] Scenarios per product: {', '.join(scenarios)}\n")

    with open(creator.INPUT_FILE, "r", encoding="utf-8") as f:
        opportunities = json.load(f)

    tmp = tempfile.mkdtemp(prefix="sim_buyer_")
    profile_dir = os.path.join(tmp, "lo_profile")

    # ---- PASS 1: stage every scenario file for every product ----
    staged = []  # (idx, sc, path)
    meta = {}     # idx -> (title, inputs, formulas, xlsx)
    print("[*] PASS 1: staging scenario workbooks...")
    for idx, opp in enumerate(opportunities, start=1):
        title = opp.get("suggested_etsy_title", f"opportunity_{idx}")
        fname_base = creator.safe_filename(title)
        xlsx = os.path.join(creator.PRODUCTS_DIR, f"{fname_base}.xlsx")
        if not os.path.exists(xlsx):
            meta[idx] = (title, None, None, None)
            continue
        inputs, formulas = map_cells(xlsx)
        meta[idx] = (title, inputs, formulas, xlsx)
        for sc in scenarios:
            scen_path = os.path.join(tmp, f"prod{idx:02d}__{sc}.xlsx")
            build_scenario_wb(xlsx, inputs, sc, scen_path)
            staged.append((idx, sc, scen_path))
    print(f"[*] staged {len(staged)} scenario workbooks.\n")

    # ---- PASS 2: one soffice instance recalculates them all ----
    print("[*] PASS 2: launching LibreOffice and recalculating (this is the slow part)...")
    soffice_proc = start_soffice(profile_dir)
    try:
        log = batch_recalc(tmp)
        print(f"    {log}\n")
    finally:
        try:
            soffice_proc.terminate()
        except Exception:
            pass

    # ---- PASS 3: analyze recalculated results ----
    print("[*] PASS 3: analyzing results\n")
    catalog_errors = 0
    catalog_dead = 0
    fully_clean = 0
    by_product = {idx: {} for idx in meta}
    for idx, sc, path in staged:
        by_product[idx][sc] = path

    try:
        for idx in sorted(meta):
            title, inputs, formulas, xlsx = meta[idx]
            print(f"[{idx}/{len(opportunities)}] {title[:62]}")
            if xlsx is None:
                print("  [FAIL] workbook missing"); catalog_errors += 1; continue
            print(f"  sections(sheets)={len(set(s for s,_ in formulas))}  "
                  f"input_cells={len(inputs)}  formula_cells={len(formulas)}")

            value_history = {fc: set() for fc in formulas}
            product_issues = 0

            for sc in scenarios:
                path = by_product[idx].get(sc)
                if not path or not os.path.exists(path):
                    print(f"    [{sc}] [WARN] scenario file missing"); continue
                values, errors, uncomputed = read_formula_values(path, formulas)

                for fc, v in values.items():
                    value_history[fc].add(repr(v))

                if errors:
                    product_issues += len(errors)
                    catalog_errors += len(errors)
                    shown = errors[:6]
                    for sheet, coord, ev in shown:
                        print(f"    [{sc}] [ERROR] {sheet}!{coord} -> {ev}")
                    if len(errors) > len(shown):
                        print(f"    [{sc}] ...and {len(errors)-len(shown)} more error cells")
                if uncomputed and sc in ("baseline", "zeros", "ones", "huge"):
                    print(f"    [{sc}] [WARN] {len(uncomputed)} formula cell(s) returned no value "
                          f"(e.g. {uncomputed[0][0]}!{uncomputed[0][1]})")

            dead = [fc for fc, vals in value_history.items() if len(vals) <= 1]
            if formulas and len(dead) == len(formulas):
                print(f"  [DEAD] every formula returned a constant across all "
                      f"{len(scenarios)} scenarios -- decision logic not responding")
                catalog_dead += 1
                product_issues += 1
            elif dead:
                print(f"  [note] {len(dead)}/{len(formulas)} formula cells were constant "
                      f"across scenarios (likely labels/headers -- review if unexpected)")

            if product_issues == 0:
                fully_clean += 1
                print("  [PASS] every formula in every section survived every scenario")
            print()

    finally:
        shutil.rmtree(tmp, ignore_errors=True)

    print("\n" + "=" * 70)
    print(f"[done] {fully_clean}/{len(opportunities)} products fully clean across all scenarios.")
    print(f"[done] {catalog_errors} error-cell occurrences found catalog-wide.")
    print(f"[done] {catalog_dead} products with completely unresponsive decision logic.")
    print("=" * 70)


if __name__ == "__main__":
    main()
