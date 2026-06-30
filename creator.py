"""
FILE 5: creator.py
Reads opportunities.json and, for each opportunity, asks Claude to design a
full dashboard spec, then builds the .xlsx (openpyxl), a companion PDF guide
(reportlab), and Etsy listing copy (.txt).
Run independently: python creator.py
Output: /products/*.xlsx, /products/*.pdf, /listings/*.txt
"""

import json
import os
import re
import sys

from anthropic import Anthropic
from dotenv import load_dotenv
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import CellIsRule
from reportlab.lib.pagesizes import letter
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, ListFlowable, ListItem

MODEL_NAME = "claude-opus-4-8"  # upgraded from sonnet-4-6 for best product quality, per user request

INPUT_FILE = "opportunities.json"
PRODUCTS_DIR = "products"
LISTINGS_DIR = "listings"

SPEC_SYSTEM_PROMPT = """You design Decision Support Systems (not templates) for small service \
businesses. Given a product opportunity, output a JSON object describing a Google Sheets / Excel \
dashboard with exactly these keys:

- ai_insight_summary: 3-5 sentences of example AI-generated insight text that would appear in a \
top banner of the sheet, explaining what the numbers mean and what action to take. Use concrete \
example phrasing like "Crew 2 generated 22% more profit than Crew 1" or "Truck 4 has exceeded \
replacement economics."
- sheets: an array of sheet objects, each with:
  - name: sheet tab name
  - columns: array of column header strings (KPI columns relevant to this business type). \
Include both raw input columns (numbers the buyer types in, e.g. "Revenue", "Labor Hours") AND \
computed columns (e.g. "Profit Margin %", "Verdict") that will be driven by live formulas.
  - formulas: array of objects, one for EVERY computed column, each with:
    - "column": the exact header text of the column this formula fills (must match an entry \
in "columns" exactly)
    - "formula_template": a real Excel formula using the bracket placeholder [Header Name] to \
reference any OTHER column in this same row by its header text (e.g. "=[Revenue]-[Materials Cost]-[Labor Cost]"). \
This is REQUIRED to be a live, working Excel formula -- not a description. Verdict/recommendation \
columns must also be live formulas using IF/nested IF/text concatenation, e.g. \
'=IF([Repair Cost Ratio]>1,"Replace - repairs exceed lease","Keep - still economical")'. \
Never invent a static text verdict; always derive it from a formula referencing other columns.
  - conditional_formatting: array of objects, each with:
    - "column": exact header text of the column to highlight
    - "operator": one of "lessThan", "greaterThan", "equal"
    - "value": the numeric threshold to compare against
    - "color": one of "RED", "YELLOW", "GREEN" (use RED for problems, YELLOW for caution, GREEN for healthy)
  - sample_rows: array of arrays containing realistic sample data matching the "columns" order, \
4-6 rows. For computed columns that will be driven by formulas, still include a plausible example \
number/text matching what the formula would produce (it will be overwritten by the live formula \
when the file is built, but helps validate your formula logic is correct).
- guide_sections: an array of objects with "heading" and "body" describing how to use the \
dashboard, what each metric means, and example AI insights the sheet surfaces -- for use in a \
companion PDF guide.

Respond with ONLY the JSON object. No prose, no markdown fences."""

LISTING_SYSTEM_PROMPT = """You write Etsy listing copy for Decision Support Systems sold to small \
service business owners. These are NOT templates -- they make decisions for the buyer by \
explaining what the data means and what action to take. Given a product opportunity, output a \
JSON object with exactly these keys:
- title: keyword-optimized Etsy title, maximum 140 characters
- description: full benefit-focused description (not feature-focused), several paragraphs, \
emphasizing the decisions this system helps the buyer make
- tags: array of exactly 13 Etsy tags. CRITICAL HARD CONSTRAINT: every tag must be 20 characters \
or fewer INCLUDING SPACES -- Etsy silently rejects any tag over 20 characters. Count the \
characters in each tag before finalizing it. Prefer 2-word tags over 3-word tags to stay under \
the limit (e.g. "bid calculator" not "contractor bid calculator pricing tool"). Do not use generic \
filler words that push a tag over budget.

The opportunity object includes a "suggested_price" field that has already been set based on \
real competitor pricing research -- use that exact price in the listing, do not invent your own.

Respond with ONLY the JSON object with keys title, description, tags. No prose, no markdown fences."""


def load_anthropic_client():
    load_dotenv()
    api_key = os.getenv("ANTHROPIC_API_KEY")
    if not api_key:
        print("[fatal] Missing ANTHROPIC_API_KEY in .env file.")
        sys.exit(1)
    try:
        return Anthropic(api_key=api_key)
    except Exception as client_err:
        print(f"[fatal] Could not initialize Anthropic client: {client_err}")
        sys.exit(1)


def sanitize_text(value):
    """Recursively replace mojibake/replacement characters Claude occasionally
    emits (e.g. a corrupted em-dash) with a plain ASCII dash, so broken glyphs
    never end up baked into a sold product."""
    if isinstance(value, str):
        return value.replace("�", "-")
    if isinstance(value, list):
        return [sanitize_text(v) for v in value]
    if isinstance(value, dict):
        return {k: sanitize_text(v) for k, v in value.items()}
    return value


def call_claude_json(client, system_prompt: str, user_prompt: str, max_tokens=4096):
    response = client.messages.create(
        model=MODEL_NAME,
        max_tokens=max_tokens,
        system=system_prompt,
        messages=[{"role": "user", "content": user_prompt}],
    )
    raw_text = "".join(block.text for block in response.content if hasattr(block, "text"))
    text = raw_text.strip()
    if text.startswith("```"):
        text = text.strip("`")
        if text.lower().startswith("json"):
            text = text[4:]
    parsed = json.loads(text.strip())
    return sanitize_text(parsed)


def safe_filename(text: str) -> str:
    cleaned = re.sub(r"[^a-zA-Z0-9_\- ]", "", text).strip()
    cleaned = re.sub(r"\s+", "_", cleaned)
    return cleaned[:60] if cleaned else "product"


INVALID_SHEET_TITLE_CHARS = re.compile(r"[\\/?*\[\]:]")


def safe_sheet_title(name: str) -> str:
    cleaned = INVALID_SHEET_TITLE_CHARS.sub("-", name).strip()
    return (cleaned or "Sheet")[:31]


EXTRA_BLANK_ROWS = 15  # rows left formula-filled but empty, for the buyer's own data

CF_COLORS = {
    "RED": "FFC7CE",
    "YELLOW": "FFEB9C",
    "GREEN": "C6EFCE",
}
CF_FONT_COLORS = {
    "RED": "9C0006",
    "YELLOW": "9C6500",
    "GREEN": "006100",
}

BRACKET_REF = re.compile(r"\[([^\]]+)\]")


def resolve_formula(template: str, header_to_letter: dict, row: int) -> str:
    """Replace [Column Header] placeholders with an actual A1-style cell ref
    for the given row, so the result is a real, live Excel formula."""
    def repl(match):
        header = match.group(1)
        letter = header_to_letter.get(header)
        return f"{letter}{row}" if letter else match.group(0)
    resolved = BRACKET_REF.sub(repl, template)
    # Guard every formula against empty/zero inputs so the ~15 blank template
    # rows (and zero-denominator cases) show BLANK instead of #DIV/0!/#VALUE!
    # to the buyer. IFERROR is the standard wrap for a fill-in-the-blanks tool.
    if resolved.startswith("=") and not resolved[1:].lstrip().upper().startswith("IFERROR("):
        resolved = f'=IFERROR({resolved[1:]},"")'
    return resolved


def build_workbook(spec: dict, opportunity: dict, path: str):
    wb = Workbook()
    wb.remove(wb.active)

    insight_text = spec.get("ai_insight_summary", "")
    used_titles = set()

    for sheet_spec in spec.get("sheets", []):
        title = safe_sheet_title(sheet_spec.get("name", "Sheet"))
        base_title, suffix = title, 1
        while title in used_titles:
            suffix += 1
            title = f"{base_title[:28]}-{suffix}"
        used_titles.add(title)
        ws = wb.create_sheet(title=title)
        columns = sheet_spec.get("columns", [])
        header_to_letter = {h: get_column_letter(i) for i, h in enumerate(columns, start=1)}
        formula_specs = sheet_spec.get("formulas", [])
        formula_by_column = {f.get("column"): f.get("formula_template", "") for f in formula_specs if f.get("column")}

        # AI insight banner
        ws.merge_cells(start_row=1, start_column=1, end_row=3, end_column=max(len(columns), 1))
        banner_cell = ws.cell(row=1, column=1, value=f"AI INSIGHT SUMMARY:\n{insight_text}")
        banner_cell.font = Font(bold=True, color="FFFFFF", size=11)
        banner_cell.fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
        banner_cell.alignment = Alignment(wrap_text=True, vertical="top")

        header_row = 5
        for col_idx, header in enumerate(columns, start=1):
            cell = ws.cell(row=header_row, column=col_idx, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            ws.column_dimensions[get_column_letter(col_idx)].width = max(16, len(str(header)) + 2)

        sample_rows = sheet_spec.get("sample_rows", [])
        total_rows = len(sample_rows) + EXTRA_BLANK_ROWS

        # Write raw (non-formula) sample values first.
        for row_offset, row_data in enumerate(sample_rows, start=1):
            for col_idx, value in enumerate(row_data, start=1):
                header = columns[col_idx - 1] if col_idx - 1 < len(columns) else None
                if header in formula_by_column:
                    continue  # computed columns are filled by the live formula below
                ws.cell(row=header_row + row_offset, column=col_idx, value=value)

        # Fill every computed column with a real, live formula for every row
        # (sample rows + extra blank rows reserved for the buyer's own data).
        for header, template in formula_by_column.items():
            col_idx = None
            for i, h in enumerate(columns, start=1):
                if h == header:
                    col_idx = i
                    break
            if col_idx is None:
                continue
            for row_offset in range(1, total_rows + 1):
                row = header_row + row_offset
                formula = resolve_formula(template, header_to_letter, row)
                ws.cell(row=row, column=col_idx, value=formula)

        # Real conditional formatting (color-coded flags), not text notes.
        for cf in sheet_spec.get("conditional_formatting", []):
            header = cf.get("column")
            operator = cf.get("operator")
            value = cf.get("value")
            color = cf.get("color", "YELLOW")
            if header not in header_to_letter or operator not in ("lessThan", "greaterThan", "equal") or value is None:
                continue
            letter = header_to_letter[header]
            cell_range = f"{letter}{header_row + 1}:{letter}{header_row + total_rows}"
            fill = PatternFill(start_color=CF_COLORS.get(color, "FFEB9C"),
                                end_color=CF_COLORS.get(color, "FFEB9C"), fill_type="solid")
            font = Font(color=CF_FONT_COLORS.get(color, "9C6500"))
            ws.conditional_formatting.add(
                cell_range,
                CellIsRule(operator=operator, formula=[str(value)], fill=fill, font=font),
            )

    if not wb.sheetnames:
        wb.create_sheet(title="Dashboard")

    wb.save(path)


def build_pdf_guide(spec: dict, opportunity: dict, path: str):
    doc = SimpleDocTemplate(path, pagesize=letter, topMargin=0.75 * inch, bottomMargin=0.75 * inch)
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle("TitleCustom", parent=styles["Title"], fontSize=20)
    heading_style = styles["Heading2"]
    body_style = styles["BodyText"]

    elements = [
        Paragraph(opportunity.get("suggested_etsy_title", "Decision Support System Guide"), title_style),
        Spacer(1, 0.3 * inch),
        Paragraph("How to Use This Dashboard", heading_style),
        Spacer(1, 0.1 * inch),
    ]

    for section in spec.get("guide_sections", []):
        elements.append(Paragraph(section.get("heading", ""), heading_style))
        elements.append(Spacer(1, 0.05 * inch))
        elements.append(Paragraph(section.get("body", ""), body_style))
        elements.append(Spacer(1, 0.2 * inch))

    elements.append(Paragraph("Example AI Insights This Sheet Surfaces", heading_style))
    elements.append(Spacer(1, 0.1 * inch))
    elements.append(Paragraph(spec.get("ai_insight_summary", ""), body_style))

    doc.build(elements)


def build_listing_txt(listing_copy: dict, opportunity: dict, path: str):
    with open(path, "w", encoding="utf-8") as f:
        f.write(f"TITLE:\n{listing_copy.get('title', '')}\n\n")
        f.write(f"PRICE: ${opportunity.get('suggested_price', '')}\n\n")
        f.write(f"DESCRIPTION:\n{listing_copy.get('description', '')}\n\n")
        f.write("TAGS:\n")
        for tag in listing_copy.get("tags", []):
            f.write(f"- {tag}\n")
        f.write(f"\nTARGET CUSTOMER: {opportunity.get('target_customer', '')}\n")
        f.write(f"OPPORTUNITY SCORE: {opportunity.get('score', '')}/10\n")


def main():
    print("=" * 60)
    print("PRODUCT CREATOR (Claude API + openpyxl + reportlab)")
    print("=" * 60)

    try:
        with open(INPUT_FILE, "r", encoding="utf-8") as f:
            opportunities = json.load(f)
    except FileNotFoundError:
        print(f"[fatal] {INPUT_FILE} not found. Run analyzer.py first.")
        sys.exit(1)
    except json.JSONDecodeError as e:
        print(f"[fatal] {INPUT_FILE} is not valid JSON: {e}")
        sys.exit(1)

    os.makedirs(PRODUCTS_DIR, exist_ok=True)
    os.makedirs(LISTINGS_DIR, exist_ok=True)

    client = load_anthropic_client()
    print(f"[*] Loaded {len(opportunities)} opportunities. Building products...\n")

    built = 0
    for idx, opportunity in enumerate(opportunities, start=1):
        title = opportunity.get("suggested_etsy_title", f"opportunity_{idx}")
        print(f"[Product {idx}/{len(opportunities)}] {title}")

        try:
            print("  [-] Requesting dashboard spec from Claude...")
            spec = call_claude_json(
                client, SPEC_SYSTEM_PROMPT,
                f"Design the dashboard for this opportunity:\n{json.dumps(opportunity, indent=2)}",
                max_tokens=8192,
            )

            print("  [-] Requesting Etsy listing copy from Claude...")
            listing_copy = call_claude_json(
                client, LISTING_SYSTEM_PROMPT,
                f"Write Etsy listing copy for this opportunity:\n{json.dumps(opportunity, indent=2)}",
                max_tokens=2048,
            )

            fname_base = safe_filename(title)
            xlsx_path = os.path.join(PRODUCTS_DIR, f"{fname_base}.xlsx")
            pdf_path = os.path.join(PRODUCTS_DIR, f"{fname_base}_guide.pdf")
            listing_path = os.path.join(LISTINGS_DIR, f"{fname_base}.txt")

            print(f"  [-] Building workbook -> {xlsx_path}")
            build_workbook(spec, opportunity, xlsx_path)

            print(f"  [-] Building PDF guide -> {pdf_path}")
            build_pdf_guide(spec, opportunity, pdf_path)

            print(f"  [-] Writing listing copy -> {listing_path}")
            build_listing_txt(listing_copy, opportunity, listing_path)

            built += 1
            print("  [+] Done.\n")

        except Exception as item_err:
            print(f"  [error] Failed to build product for '{title}': {item_err}\n")
            continue

    print("=" * 60)
    print(f"[done] Built {built}/{len(opportunities)} products.")
    print(f"[done] Files saved to /{PRODUCTS_DIR} and /{LISTINGS_DIR}")
    print("=" * 60)


if __name__ == "__main__":
    main()
