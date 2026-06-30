"""
buyer_test.py  --  "Dummy buyer" end-to-end test.

Unlike stress_test.py (which only reads formula strings statically), this
actually OPENS every product workbook in a real spreadsheet engine
(LibreOffice headless), forces a full recalculation, and reports any error
cells a real buyer would see: #DIV/0!, #REF!, #VALUE!, #NAME?, #N/A, etc.

It does this by converting each .xlsx to a recalculated copy via LibreOffice's
headless "convert-to xlsx" (which recalculates on load when configured), then
reading the resulting cached values with openpyxl (data_only=True) and scanning
for Excel error strings.

Run: python buyer_test.py
Requires LibreOffice installed (soffice.exe).
"""
import glob
import json
import os
import shutil
import subprocess
import sys
import tempfile

from openpyxl import load_workbook

import creator

ERROR_VALUES = {
    "#DIV/0!", "#REF!", "#VALUE!", "#NAME?", "#N/A",
    "#NULL!", "#NUM!", "#ERROR!", "#CALC!", "Err:502",
}

SOFFICE_CANDIDATES = [
    r"C:\Program Files\LibreOffice\program\soffice.exe",
    r"C:\Program Files (x86)\LibreOffice\program\soffice.exe",
]


def find_soffice():
    for p in SOFFICE_CANDIDATES:
        if os.path.exists(p):
            return p
    found = shutil.which("soffice")
    return found


# LibreOffice will NOT recalculate OOXML files on load unless told to. Our
# openpyxl-built files have no cached formula values, so without forcing recalc
# every formula cell would read back as None and errors would never surface.
# This profile config sets OOXMLRecalcMode + ODFRecalcMode = 0 ("Always recalc").
RECALC_XCU = """<?xml version="1.0" encoding="UTF-8"?>
<oor:items xmlns:oor="http://openoffice.org/2001/registry" xmlns:xs="http://www.w3.org/2001/XMLSchema" xmlns:xsi="http://www.w3.org/2001/XMLSchema-instance">
 <item oor:path="/org.openoffice.Office.Calc/Formula/Load"><prop oor:name="OOXMLRecalcMode" oor:op="fuse"><value>0</value></prop></item>
 <item oor:path="/org.openoffice.Office.Calc/Formula/Load"><prop oor:name="ODFRecalcMode" oor:op="fuse"><value>0</value></prop></item>
</oor:items>
"""


def make_profile(base_tmp):
    """Build a throwaway LO user profile that forces full recalculation."""
    profile = os.path.join(base_tmp, "lo_profile")
    user_cfg = os.path.join(profile, "user")
    os.makedirs(user_cfg, exist_ok=True)
    with open(os.path.join(user_cfg, "registrymodifications.xcu"), "w", encoding="utf-8") as f:
        f.write(RECALC_XCU)
    # LibreOffice wants a file:// URL for -env:UserInstallation
    return "file:///" + profile.replace("\\", "/")


def recalc_to_temp(soffice, xlsx_path, outdir, profile_url):
    """Open in LibreOffice headless, force-recalc, re-save as xlsx with cached
    values. Returns recalced path."""
    cmd = [
        soffice, f"-env:UserInstallation={profile_url}",
        "--headless", "--calc",
        "--convert-to", "xlsx:Calc MS Excel 2007 XML",
        "--outdir", outdir, xlsx_path,
    ]
    subprocess.run(cmd, capture_output=True, timeout=120,
                   env={**os.environ, "HOME": outdir})
    base = os.path.splitext(os.path.basename(xlsx_path))[0] + ".xlsx"
    return os.path.join(outdir, base)


def scan_errors(recalced_path):
    issues = []
    try:
        wb = load_workbook(recalced_path, data_only=True)
    except Exception as e:
        return [f"could not reopen recalculated workbook: {e}"]
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                v = cell.value
                if isinstance(v, str) and v.strip() in ERROR_VALUES:
                    issues.append(f"sheet '{ws.title}' cell {cell.coordinate}: buyer sees {v.strip()}")
    return issues


def main():
    soffice = find_soffice()
    if not soffice:
        print("[FATAL] LibreOffice (soffice.exe) not found. Install it, then re-run.")
        sys.exit(2)
    print(f"[*] Using LibreOffice: {soffice}\n")

    with open(creator.INPUT_FILE, "r", encoding="utf-8") as f:
        opportunities = json.load(f)

    tmp = tempfile.mkdtemp(prefix="buyer_test_")
    profile_url = make_profile(tmp)
    total_err = 0
    clean = 0
    try:
        for idx, opp in enumerate(opportunities, start=1):
            title = opp.get("suggested_etsy_title", f"opportunity_{idx}")
            fname_base = creator.safe_filename(title)
            xlsx = os.path.join(creator.PRODUCTS_DIR, f"{fname_base}.xlsx")
            print(f"[{idx}/{len(opportunities)}] {title[:65]}")
            if not os.path.exists(xlsx):
                print("  [FAIL] workbook missing")
                total_err += 1
                continue
            try:
                recalced = recalc_to_temp(soffice, xlsx, tmp, profile_url)
                if not os.path.exists(recalced):
                    print("  [WARN] LibreOffice produced no output (open it manually to check)")
                    continue
                errs = scan_errors(recalced)
            except subprocess.TimeoutExpired:
                print("  [FAIL] LibreOffice timed out opening this file")
                total_err += 1
                continue
            if errs:
                total_err += len(errs)
                for e in errs:
                    print(f"  [FAIL] {e}")
            else:
                clean += 1
                print("  [PASS] opens clean, no error cells on recalculation")
    finally:
        shutil.rmtree(tmp, ignore_errors=True)

    print("\n" + "=" * 68)
    print(f"[done] {clean}/{len(opportunities)} workbooks recalculated with ZERO error cells.")
    print(f"[done] {total_err} error cells/failures found across the catalog.")
    print("=" * 68)


if __name__ == "__main__":
    main()
