"""One-off fix: wrap every formula in every already-built product workbook in
IFERROR(...,"") so empty template rows and zero-denominator cases show BLANK
instead of #DIV/0! / #VALUE! to the buyer. Costs zero API tokens -- it edits the
existing .xlsx files in place, preserving the exact formulas otherwise."""
import json
import os

from openpyxl import load_workbook

import creator

with open(creator.INPUT_FILE, "r", encoding="utf-8") as f:
    opportunities = json.load(f)

patched_files = 0
patched_cells = 0
for idx, opp in enumerate(opportunities, start=1):
    title = opp.get("suggested_etsy_title", f"opportunity_{idx}")
    path = os.path.join(creator.PRODUCTS_DIR, f"{creator.safe_filename(title)}.xlsx")
    if not os.path.exists(path):
        print(f"[skip] missing: {title[:50]}")
        continue
    wb = load_workbook(path)
    changed = 0
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                v = cell.value
                if isinstance(v, str) and v.startswith("=") \
                        and not v[1:].lstrip().upper().startswith("IFERROR("):
                    cell.value = f'=IFERROR({v[1:]},"")'
                    changed += 1
    if changed:
        wb.save(path)
        patched_files += 1
        patched_cells += changed
    wb.close()
    print(f"[{idx}/{len(opportunities)}] wrapped {changed:4d} formulas | {title[:48]}")

print(f"\n[done] patched {patched_cells} formula cells across {patched_files} workbooks.")
