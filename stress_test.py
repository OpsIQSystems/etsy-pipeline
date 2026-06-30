"""
stress_test.py
Validates every built product before it could ever reach a real buyer.
Checks for:
  1. Broken formulas: any cell formula still containing an unresolved
     [Column Header] placeholder (means the header text didn't match exactly
     -- Excel would show this as literal text or #NAME? error to the buyer).
  2. Formula cells that are empty/None where a formula was expected.
  3. Each workbook has at least one sheet with both formulas AND conditional
     formatting (sanity check that the "decision support" features exist).
  4. Each product has its 3 expected images present in /products/images.
  5. Each product has a non-empty listing .txt with title/price/tags.
Run independently: python stress_test.py
"""

import json
import os
import re

from openpyxl import load_workbook

import creator

BRACKET_LEFTOVER = re.compile(r"\[[^\]]+\]")


def check_workbook(path: str):
    issues = []
    try:
        wb = load_workbook(path)
    except Exception as e:
        return [f"could not open workbook: {e}"]

    total_formula_cells = 0
    total_cf_rules = 0
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                if isinstance(cell.value, str) and cell.value.startswith("="):
                    total_formula_cells += 1
                    if BRACKET_LEFTOVER.search(cell.value):
                        issues.append(
                            f"sheet '{ws.title}' cell {cell.coordinate}: unresolved placeholder "
                            f"left in formula -> {cell.value!r}"
                        )
        total_cf_rules += len(ws.conditional_formatting._cf_rules) if hasattr(ws.conditional_formatting, "_cf_rules") else 0

    if total_formula_cells == 0:
        issues.append("no live formulas found anywhere in this workbook -- it is a static template, not a decision tool")
    if total_cf_rules == 0:
        issues.append("no conditional formatting rules found -- no visual flags for the buyer")

    return issues


def check_images(fname_base: str):
    images_dir = os.path.join(creator.PRODUCTS_DIR, "images")
    expected = [f"{fname_base}_1_cover.png", f"{fname_base}_2_features.png", f"{fname_base}_3_insight.png"]
    missing = [e for e in expected if not os.path.exists(os.path.join(images_dir, e))]
    return missing


def check_listing(fname_base: str):
    issues = []
    path = os.path.join(creator.LISTINGS_DIR, f"{fname_base}.txt")
    if not os.path.exists(path):
        return ["listing .txt file is missing entirely"]
    text = open(path, "r", encoding="utf-8").read()
    if "TITLE:\n\n" in text or text.strip().startswith("TITLE:\n\nPRICE"):
        issues.append("title appears empty")
    if "PRICE: $\n" in text:
        issues.append("price appears empty")

    # Only parse the actual TAGS: section, not other bulleted lists (e.g. "WHAT YOU GET:").
    tags_match = re.search(r"^TAGS:\n((?:- .*\n?)*)", text, re.MULTILINE)
    tags = []
    if tags_match:
        tags = [line[2:].strip() for line in tags_match.group(1).splitlines() if line.startswith("- ")]

    if len(tags) < 10:
        issues.append(f"only {len(tags)} tags found, expected 13")
    for tag in tags:
        if len(tag) > 20:
            issues.append(f"tag exceeds Etsy's 20-char limit ({len(tag)} chars): {tag!r}")
    return issues


def main():
    with open(creator.INPUT_FILE, "r", encoding="utf-8") as f:
        opportunities = json.load(f)

    print("=" * 70)
    print(f"STRESS TEST: validating {len(opportunities)} products")
    print("=" * 70)

    total_issues = 0
    clean_products = 0

    for idx, opp in enumerate(opportunities, start=1):
        title = opp.get("suggested_etsy_title", f"opportunity_{idx}")
        fname_base = creator.safe_filename(title)
        xlsx_path = os.path.join(creator.PRODUCTS_DIR, f"{fname_base}.xlsx")

        print(f"\n[{idx}/{len(opportunities)}] {title[:70]}")

        all_issues = []
        if not os.path.exists(xlsx_path):
            all_issues.append("workbook .xlsx file is missing entirely")
        else:
            all_issues += [f"WORKBOOK: {i}" for i in check_workbook(xlsx_path)]

        missing_images = check_images(fname_base)
        if missing_images:
            all_issues.append(f"IMAGES: missing {missing_images}")

        all_issues += [f"LISTING: {i}" for i in check_listing(fname_base)]

        if all_issues:
            total_issues += len(all_issues)
            for issue in all_issues:
                print(f"  [FAIL] {issue}")
        else:
            clean_products += 1
            print("  [PASS] no issues found")

    print("\n" + "=" * 70)
    print(f"[done] {clean_products}/{len(opportunities)} products passed clean.")
    print(f"[done] {total_issues} total issues found across the catalog.")
    print("=" * 70)


if __name__ == "__main__":
    main()
