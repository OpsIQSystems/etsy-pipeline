"""
enhance_commercial.py

Builds the underwriter's #1 missing piece directly INTO the Commercial Real
Estate Deal Analyzer (instead of a disclaimer):

  * two new input columns -- Capital Reserve $/SqFt and Management Fee % --
    placed with the other inputs, and
  * NOI now deducts a replacement/CapEx reserve and a management fee (the two
    non-reimbursable line items a real credit committee always loads in), and
  * break-even occupancy now includes those costs, and
  * the verdict closes the value-add gap: a high-vacancy deal (>15%) can no
    longer earn an unconditional "Buy" just because its goal is "Appreciation".

The Deal Comparison sheet's computed region is rebuilt deterministically with a
clean left-to-right layout (inputs, then computed), preserving the banner,
header styling, and conditional-formatting cues.
"""
import os
from copy import copy

from openpyxl import load_workbook
from openpyxl.formatting.formatting import ConditionalFormattingList
from openpyxl.utils import get_column_letter

import creator
from build_commercial import COMMERCIAL

FNAME = creator.safe_filename(COMMERCIAL["suggested_etsy_title"])
XLSX = os.path.join(creator.PRODUCTS_DIR, f"{FNAME}.xlsx")
SHEET = "Deal Comparison"
HDR_ROW = 5
FIRST = 6
LAST = 12          # formula rows run 6..12 (sample data in 6..10)

# ---- new column layout (letter -> header) -------------------------------
INPUT_HEADERS = [
    ("A", "Property Name"),
    ("B", "Asset Type"),
    ("C", "Square Footage"),
    ("D", "Purchase Price"),
    ("E", "Gross Rental Income"),
    ("F", "Vacancy %"),
    ("G", "Operating Expenses"),
    ("H", "Capital Reserve $/SqFt"),     # NEW
    ("I", "Management Fee %"),           # NEW
    ("J", "Lease Type"),
    ("K", "Loan Amount"),
    ("L", "Interest Rate"),
    ("M", "Amortization Years"),
    ("N", "Target DSCR"),
    ("O", "Benchmark Price Per SqFt"),
    ("P", "Investor Goal"),
]
COMPUTED = [
    ("Q", "Effective Gross Income", "=IFERROR(E{r}*(1-F{r}),\"\")"),
    ("R", "Landlord Op Ex Responsibility",
     "=IFERROR(IF(J{r}=\"NNN\",G{r}*0.05,IF(J{r}=\"Modified Gross\",G{r}*0.5,G{r})),\"\")"),
    ("S", "Capital Reserve", "=IFERROR(H{r}*C{r},\"\")"),
    ("T", "Management Fee", "=IFERROR(I{r}*Q{r},\"\")"),
    ("U", "NOI", "=IFERROR(Q{r}-R{r}-S{r}-T{r},\"\")"),
    ("V", "Cap Rate %", "=IFERROR(U{r}/D{r},\"\")"),
    ("W", "Price Per SqFt", "=IFERROR(D{r}/C{r},\"\")"),
    ("X", "Price vs Benchmark %", "=IFERROR((W{r}-O{r})/O{r},\"\")"),
    ("Y", "Loan Constant %", "=IFERROR(((L{r}/12)/(1-(1+L{r}/12)^(-M{r}*12)))*12,\"\")"),
    ("Z", "Annual Debt Service", "=IFERROR(K{r}*Y{r},\"\")"),
    ("AA", "DSCR", "=IFERROR(U{r}/Z{r},\"\")"),
    ("AB", "Cash Invested", "=IFERROR(D{r}-K{r},\"\")"),
    ("AC", "Cash-on-Cash Return %", "=IFERROR((U{r}-Z{r})/AB{r},\"\")"),
    ("AD", "Break-Even Occupancy %", "=IFERROR((R{r}+S{r}+T{r}+Z{r})/E{r},\"\")"),
    ("AE", "Verdict",
     "=IFERROR(IF(OR(AA{r}<N{r},AD{r}>0.95),"
     "IF(X{r}>0.05,\"Pass - underwater on DSCR and overpriced\","
     "\"Negotiate - financing too tight at this price\"),"
     "IF(X{r}>0.05,\"Negotiate - price above submarket benchmark\","
     "IF(F{r}>0.15,\"Negotiate - high vacancy, underwrite lease-up first\","
     "IF(AND(P{r}=\"Cash Flow\",AC{r}<0.06),\"Negotiate - cash flow below goal target\","
     "\"Buy - meets lender and goal thresholds\")))),\"\")"),
]
LAST_COL = "AE"  # 31

# Input field order matching the sample-deal dicts (without the two new ones)
SAMPLE = [
    dict(property_name="Maple Plaza Retail", asset_type="Retail", square_footage=12000,
         purchase_price=2100000, gross_rental_income=198000, vacancy_pct=0.05,
         operating_expenses=42000, lease_type="NNN", loan_amount=1470000,
         interest_rate=0.072, amortization_years=25, target_dscr=1.25,
         benchmark_price_per_sqft=191, investor_goal="Cash Flow"),
    dict(property_name="Riverside Office", asset_type="Office", square_footage=15500,
         purchase_price=2350000, gross_rental_income=212000, vacancy_pct=0.08,
         operating_expenses=68000, lease_type="Modified Gross", loan_amount=1645000,
         interest_rate=0.075, amortization_years=25, target_dscr=1.25,
         benchmark_price_per_sqft=191, investor_goal="Appreciation"),
    dict(property_name="Industrial Park Unit C", asset_type="Industrial", square_footage=22000,
         purchase_price=1850000, gross_rental_income=168000, vacancy_pct=0.04,
         operating_expenses=38000, lease_type="NNN", loan_amount=1295000,
         interest_rate=0.069, amortization_years=25, target_dscr=1.25,
         benchmark_price_per_sqft=88, investor_goal="Cash Flow"),
    dict(property_name="Downtown Mixed-Use", asset_type="Mixed-Use", square_footage=9800,
         purchase_price=1650000, gross_rental_income=156000, vacancy_pct=0.07,
         operating_expenses=71000, lease_type="Full-Service Gross", loan_amount=1155000,
         interest_rate=0.074, amortization_years=25, target_dscr=1.3,
         benchmark_price_per_sqft=175, investor_goal="Appreciation"),
    dict(property_name="Eastgate Strip Center", asset_type="Retail", square_footage=14200,
         purchase_price=1980000, gross_rental_income=176000, vacancy_pct=0.06,
         operating_expenses=40000, lease_type="NNN", loan_amount=1386000,
         interest_rate=0.071, amortization_years=25, target_dscr=1.25,
         benchmark_price_per_sqft=155, investor_goal="Cash Flow"),
]
# defensible reserve/management defaults by asset class
RESERVE_PSF = {"Retail": 0.20, "Office": 0.25, "Industrial": 0.15, "Mixed-Use": 0.25}
MGMT_PCT = {"Retail": 0.04, "Office": 0.04, "Industrial": 0.03, "Mixed-Use": 0.05}

INPUT_KEY_BY_COL = {
    "A": "property_name", "B": "asset_type", "C": "square_footage", "D": "purchase_price",
    "E": "gross_rental_income", "F": "vacancy_pct", "G": "operating_expenses",
    "H": "capital_reserve_per_sqft", "I": "management_fee_pct", "J": "lease_type",
    "K": "loan_amount", "L": "interest_rate", "M": "amortization_years", "N": "target_dscr",
    "O": "benchmark_price_per_sqft", "P": "investor_goal",
}
WIDTHS = {"A": 18, "B": 14, "C": 15, "D": 15, "E": 19, "F": 12, "G": 18, "H": 21, "I": 17,
          "J": 18, "K": 14, "L": 14, "M": 18, "N": 13, "O": 24, "P": 15}

# old computed col -> new computed col, to carry conditional-format cues over
CF_REMAP = {"R": "V", "T": "X", "W": "AA", "Y": "AC", "Z": "AD"}


def main():
    wb = load_workbook(XLSX)
    ws = wb[SHEET]

    # capture header styling + existing conditional-format rules (with their dxf)
    hdr = ws["A5"]
    hfill, hfont, halign, hborder = copy(hdr.fill), copy(hdr.font), copy(hdr.alignment), copy(hdr.border)
    old_rules = []
    for rng, rules in list(ws.conditional_formatting._cf_rules.items()):
        col = str(rng.sqref).split(":")[0].rstrip("0123456789").lstrip("$")
        for rule in rules:
            old_rules.append((col, rule))

    # fill in reserve/mgmt on the sample deals
    deals = []
    for d in SAMPLE:
        d = dict(d)
        d["capital_reserve_per_sqft"] = RESERVE_PSF.get(d["asset_type"], 0.20)
        d["management_fee_pct"] = MGMT_PCT.get(d["asset_type"], 0.04)
        deals.append(d)

    # ---- headers ----
    for col, label in INPUT_HEADERS + [(c, h) for c, h, _ in COMPUTED]:
        cell = ws[f"{col}{HDR_ROW}"]
        cell.value = label
        cell.fill, cell.font, cell.alignment, cell.border = copy(hfill), copy(hfont), copy(halign), copy(hborder)

    # ---- input cells (rows 6..LAST; sample data in first len(deals) rows) ----
    for i in range(FIRST, LAST + 1):
        deal = deals[i - FIRST] if (i - FIRST) < len(deals) else None
        for col, key in INPUT_KEY_BY_COL.items():
            ws[f"{col}{i}"] = deal[key] if deal else None

    # ---- computed formulas (every row 6..LAST so blank rows stay ready) ----
    for col, _, tmpl in COMPUTED:
        for i in range(FIRST, LAST + 1):
            ws[f"{col}{i}"] = tmpl.format(r=i)

    # ---- widths ----
    for col, w in WIDTHS.items():
        ws.column_dimensions[col].width = w

    # ---- re-merge the AI insight banner across the wider sheet ----
    for mr in list(ws.merged_cells.ranges):
        if mr.min_row == 1:
            ws.unmerge_cells(str(mr))
    ws.merge_cells(f"A1:{LAST_COL}3")

    # ---- carry conditional-format cues onto the new computed columns ----
    ws.conditional_formatting = ConditionalFormattingList()
    for old_col, rule in old_rules:
        new_col = CF_REMAP.get(old_col)
        if not new_col:
            continue
        ws.conditional_formatting.add(f"{new_col}{FIRST}:{new_col}25", rule)

    wb.save(XLSX)
    wb.close()
    print(f"[done] rebuilt model -> {XLSX}")
    print("       new inputs: Capital Reserve $/SqFt (H), Management Fee % (I)")
    print("       NOI now = EGI - Landlord OpEx - Reserve - Mgmt Fee")
    print("       verdict now gates high-vacancy (>15%) deals regardless of goal")


if __name__ == "__main__":
    main()
